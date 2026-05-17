"""
Excel Export Service — Generates downloadable .xlsx workbooks from tender results.

Sheets:
  1. "BOQ Items"     — Extracted Bill of Quantities with pricing and confidence
  2. "Pricing Summary" — Pricing breakdown (labour, materials, transport, VAT, etc.)
  3. "Workforce"     — Workforce requirements with categories and confidence
  4. "Warnings"      — Pipeline warnings, failed stages, retry metadata, extraction notes

HONESTY RULES:
  - If BOQ is unavailable: Sheet states "Unavailable due to failed processing stage"
  - If pricing is unavailable: Sheet states "Unavailable due to failed processing stage"
  - If workforce is unavailable: Sheet states "Unavailable due to failed processing stage"
  - NO fake data, NO placeholders, NO fabricated values
  - Confidence levels preserved (High / Medium / Low)
  - Failed stages visible with reasons
  - partial_success distinctions maintained
"""
import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")

DATA_FONT = Font(name="Calibri", size=10)
WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
UNAVAILABLE_FILL = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")
CURRENCY_FORMAT = '#,##0.00'
NUMBER_FORMAT = '#,##0'
PERCENTAGE_FORMAT = '0%'


def _style_header_row(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def _style_data_cell(cell, is_currency: bool = False, is_number: bool = False):
    """Apply standard data cell styling."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGNMENT
    if is_currency:
        cell.number_format = CURRENCY_FORMAT
    elif is_number:
        cell.number_format = NUMBER_FORMAT


def _auto_column_width(ws, col_count: int, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 12)


def _write_unavailable_sheet(ws, title: str, reason: str):
    """Write a standardized 'unavailable' message to a sheet."""
    ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=3, column=1, value=f"Unavailable due to failed processing stage").font = Font(
        name="Calibri", size=11, italic=True, color="666666"
    )
    ws.cell(row=4, column=1, value=reason).font = Font(
        name="Calibri", size=10, color="999999"
    )
    ws.cell(row=4, column=1).fill = UNAVAILABLE_FILL
    ws.column_dimensions["A"].width = 60


def _get_pricing_confidence(pricing_result: Optional[Dict[str, Any]]) -> str:
    """Extract confidence level from pricing result dict."""
    if not pricing_result:
        return "N/A"
    return pricing_result.get("confidence") or pricing_result.get("price_reliability") or "Estimated"


def generate_export(job_id: str, result_data: Dict[str, Any]) -> BytesIO:
    """
    Generate a complete Excel workbook from a processing result.

    Args:
        job_id: The job ID for filename generation
        result_data: The full ProcessingResult as a dict (from result_json)

    Returns:
        BytesIO stream containing the .xlsx workbook
    """
    wb = Workbook()
    logger.info("[EXPORT] Generating Excel export for job %s", job_id)

    status = result_data.get("status", "unknown")
    filename = result_data.get("filename", "Unknown")
    completed_stages = result_data.get("completed_stages", [])
    failed_stages = result_data.get("failed_stages", [])
    warnings = result_data.get("warnings", [])

    # ── Sheet 1: BOQ Items ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "BOQ Items"
    boq_items = result_data.get("boq_items", [])

    if not boq_items:
        _write_unavailable_sheet(
            ws1,
            "Bill of Quantities",
            f"BOQ extraction stage status: "
            f"{'failed' if 'boq_analysis' in failed_stages else 'not_available'}.\n"
            f"No BOQ items were extracted during processing.",
        )
    else:
        # Title
        ws1.cell(row=1, column=1, value=f"Bill of Quantities — {filename}").font = Font(
            name="Calibri", size=14, bold=True
        )
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        # Status indicator
        status_row = 2
        ws1.cell(row=status_row, column=1, value=f"Status: {status}").font = Font(
            name="Calibri", size=10, italic=True
        )

        # Headers
        headers = ["Item Number", "Description", "Quantity", "Unit", "Rate (ZAR)", "Amount (ZAR)", "Confidence", "Notes"]
        header_row = 4
        for col, header in enumerate(headers, 1):
            ws1.cell(row=header_row, column=col, value=header)
        _style_header_row(ws1, header_row, len(headers))

        # Data rows
        boq_confidence = result_data.get("boq_confidence", "N/A")
        for i, item in enumerate(boq_items):
            row = header_row + 1 + i
            ws1.cell(row=row, column=1, value=item.get("item_no", ""))
            ws1.cell(row=row, column=2, value=item.get("description", ""))
            ws1.cell(row=row, column=3, value=item.get("quantity"))
            ws1.cell(row=row, column=4, value=item.get("unit", ""))
            ws1.cell(row=row, column=5, value=item.get("rate"))
            ws1.cell(row=row, column=6, value=item.get("amount"))
            ws1.cell(row=row, column=7, value=boq_confidence)

            # Notes column — mark any items with low/no pricing data
            notes = []
            if item.get("rate") is None or item.get("amount") is None:
                notes.append("No rate/amount extracted")
            ws1.cell(row=row, column=8, value="; ".join(notes) if notes else "")

            # Style cells
            for col in range(1, len(headers) + 1):
                cell = ws1.cell(row=row, column=col)
                is_curr = col in (5, 6)
                is_num = col == 3
                _style_data_cell(cell, is_currency=is_curr, is_number=is_num)

        # Summary row
        summary_row = header_row + 1 + len(boq_items) + 1
        ws1.cell(row=summary_row, column=1, value="Total Items").font = Font(
            name="Calibri", size=10, bold=True
        )
        ws1.cell(row=summary_row, column=2, value=len(boq_items))
        ws1.cell(row=summary_row, column=1).fill = SUBHEADER_FILL
        ws1.cell(row=summary_row, column=2).fill = SUBHEADER_FILL

        # Add confidence note
        note_row = summary_row + 1
        ws1.cell(row=note_row, column=1, value=f"BOQ Confidence: {boq_confidence}").font = Font(
            name="Calibri", size=9, italic=True, color="666666"
        )

        _auto_column_width(ws1, len(headers))

        logger.info("[EXPORT] Added %d BOQ rows for job %s", len(boq_items), job_id)

    # ── Sheet 2: Pricing Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("Pricing Summary")
    pricing_result = result_data.get("pricing_result")
    pricing_status = result_data.get("pricing_status")

    if not pricing_result or pricing_status == "failed" or not pricing_result:
        _write_unavailable_sheet(
            ws2,
            "Pricing Summary",
            f"Pricing calculation stage status: "
            f"{'failed' if 'pricing_calculation' in failed_stages else 'not_available'}.\n"
            f"Reason: {result_data.get('pricing_unavailable_reason', 'Unknown')}",
        )
    else:
        ws2.cell(row=1, column=1, value=f"Pricing Summary — {filename}").font = Font(
            name="Calibri", size=14, bold=True
        )
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        # Pricing confidence
        conf = _get_pricing_confidence(pricing_result)
        ws2.cell(row=2, column=1, value=f"Pricing Confidence: {conf}").font = Font(
            name="Calibri", size=10, italic=True
        )
        ws2.cell(row=3, column=1, value=f"Calculation Method: {pricing_result.get('price_reliability', 'N/A')}").font = Font(
            name="Calibri", size=10, italic=True
        )

        # Pricing breakdown
        pricing_fields = [
            ("Labour Cost", "labour_cost"),
            ("Materials Cost", "materials_cost"),
            ("Transport Cost", "transport_cost"),
            ("Overheads", "overheads"),
            ("Subtotal", "subtotal"),
            ("VAT", "vat"),
            ("Total Monthly", "total_monthly"),
            ("Total Annual", "total_annual"),
            ("Final Contract Value", "final_contract_value"),
        ]

        field_row = 5
        ws2.cell(row=field_row, column=1, value="Component")
        ws2.cell(row=field_row, column=2, value="Amount (ZAR)")
        ws2.cell(row=field_row, column=3, value="Notes")
        _style_header_row(ws2, field_row, 3)

        row_idx = field_row + 1
        for label, key in pricing_fields:
            value = pricing_result.get(key)
            ws2.cell(row=row_idx, column=1, value=label)
            if value is not None:
                ws2.cell(row=row_idx, column=2, value=value)
            else:
                ws2.cell(row=row_idx, column=2, value="—")
            ws2.cell(row=row_idx, column=3, value="")
            for col in range(1, 4):
                cell = ws2.cell(row=row_idx, column=col)
                is_curr = col == 2
                _style_data_cell(cell, is_currency=is_curr)
            row_idx += 1

        # Note about estimation
        note_row = row_idx + 1
        if pricing_result.get("price_reliability") == "boq_based":
            ws2.cell(row=note_row, column=1, value="Pricing based on extracted Bill of Quantities.").font = Font(
                name="Calibri", size=9, italic=True, color="1F4E79"
            )
        elif pricing_result.get("price_reliability") == "estimated":
            ws2.cell(row=note_row, column=1, value="Pricing is estimated (no BOQ data available).").font = Font(
                name="Calibri", size=9, italic=True, color="856404"
            )
            ws2.cell(row=note_row, column=1).fill = WARN_FILL
        elif pricing_result.get("price_reliability") == "low":
            ws2.cell(row=note_row, column=1, value="Low confidence pricing — BOQ extracted with reduced reliability.").font = Font(
                name="Calibri", size=9, italic=True, color="856404"
            )
            ws2.cell(row=note_row, column=1).fill = WARN_FILL

        _auto_column_width(ws2, 3)

    # ── Sheet 3: Workforce ───────────────────────────────────────────
    ws3 = wb.create_sheet("Workforce")
    workforce_data = result_data.get("detected_workforce", {})

    if not workforce_data:
        _write_unavailable_sheet(
            ws3,
            "Workforce Analysis",
            f"Entity extraction status: "
            f"{'failed' if 'entity_extraction' in failed_stages else 'not_available'}.\n"
            f"No workforce data was extracted during processing.",
        )
    else:
        ws3.cell(row=1, column=1, value=f"Workforce Analysis — {filename}").font = Font(
            name="Calibri", size=14, bold=True
        )
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        # Workforce confidence
        wf_conf = workforce_data.get("workforce_inference_confidence") or \
                  result_data.get("boq_confidence", "N/A")
        ws3.cell(row=2, column=1, value=f"Workforce Confidence: {wf_conf}").font = Font(
            name="Calibri", size=10, italic=True
        )

        field_row = 4
        ws3.cell(row=field_row, column=1, value="Category")
        ws3.cell(row=field_row, column=2, value="Count")
        ws3.cell(row=field_row, column=3, value="Notes")
        _style_header_row(ws3, field_row, 3)

        workforce_fields = [
            ("Skilled Workers", "skilled_workers"),
            ("Unskilled Workers", "unskilled_workers"),
            ("Supervisors", "supervisors"),
            ("Total Workers", "total_workers"),
            ("Shifts Per Day", "shifts_per_day"),
            ("Hours Per Day", "hours_per_day"),
            ("Days Per Week", "days_per_week"),
        ]

        row_idx = field_row + 1
        for label, key in workforce_fields:
            value = workforce_data.get(key)
            ws3.cell(row=row_idx, column=1, value=label)
            if value is not None:
                ws3.cell(row=row_idx, column=2, value=value)
            else:
                ws3.cell(row=row_idx, column=2, value="—")
            ws3.cell(row=row_idx, column=3, value="")
            for col in range(1, 4):
                cell = ws3.cell(row=row_idx, column=col)
                _style_data_cell(cell, is_number=(col == 2))
            row_idx += 1

        # Work categories if available
        work_categories = workforce_data.get("work_categories", [])
        if work_categories:
            cat_row = row_idx + 1
            ws3.cell(row=cat_row, column=1, value="Work Categories").font = SUBHEADER_FONT
            ws3.cell(row=cat_row, column=1).fill = SUBHEADER_FILL
            cat_row += 1
            for cat in work_categories:
                ws3.cell(row=cat_row, column=1, value=cat)
                ws3.cell(row=cat_row, column=1).font = DATA_FONT
                cat_row += 1

        # Reasoning note
        if workforce_data.get("workforce_reasoning"):
            reason_row = row_idx + 2
            ws3.cell(row=reason_row, column=1, value="Notes").font = SUBHEADER_FONT
            ws3.cell(row=reason_row, column=1).fill = SUBHEADER_FILL
            ws3.cell(row=reason_row + 1, column=1, value=workforce_data["workforce_reasoning"]).font = Font(
                name="Calibri", size=9, italic=True, color="666666"
            )

        _auto_column_width(ws3, 3)

    # ── Sheet 4: Warnings ────────────────────────────────────────────
    ws4 = wb.create_sheet("Warnings")
    ws4.cell(row=1, column=1, value=f"Processing Warnings — {filename}").font = Font(
        name="Calibri", size=14, bold=True
    )
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    ws4.cell(row=2, column=1, value=f"Final Status: {status}").font = Font(
        name="Calibri", size=10, bold=True
    )

    # Completion info
    info_row = 4
    ws4.cell(row=info_row, column=1, value="Completed Stages").font = SUBHEADER_FONT
    ws4.cell(row=info_row, column=1).fill = SUBHEADER_FILL
    if completed_stages:
        for i, stage in enumerate(completed_stages):
            row = info_row + 1 + i
            ws4.cell(row=row, column=1, value=stage.replace("_", " ").title())
            ws4.cell(row=row, column=1).fill = SUCCESS_FILL
            ws4.cell(row=row, column=1).font = DATA_FONT

    failed_info_row = info_row + max(len(completed_stages), 1) + 1
    ws4.cell(row=failed_info_row, column=1, value="Failed Stages").font = SUBHEADER_FONT
    ws4.cell(row=failed_info_row, column=1).fill = SUBHEADER_FILL
    if failed_stages:
        for i, stage in enumerate(failed_stages):
            row = failed_info_row + 1 + i
            ws4.cell(row=row, column=1, value=stage.replace("_", " ").title())
            ws4.cell(row=row, column=1).fill = ERROR_FILL
            ws4.cell(row=row, column=1).font = DATA_FONT
    else:
        row = failed_info_row + 1
        ws4.cell(row=row, column=1, value="No stages failed").font = Font(
            name="Calibri", size=10, italic=True, color="28A745"
        )

    # Warnings list
    warn_start_row = failed_info_row + max(len(failed_stages), 1) + 2
    ws4.cell(row=warn_start_row, column=1, value="Warnings & Notes").font = SUBHEADER_FONT
    ws4.cell(row=warn_start_row, column=1).fill = SUBHEADER_FILL
    if warnings:
        for i, warning in enumerate(warnings):
            row = warn_start_row + 1 + i
            ws4.cell(row=row, column=1, value=warning)
            ws4.cell(row=row, column=1).font = DATA_FONT
            ws4.cell(row=row, column=1).fill = WARN_FILL
            ws4.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    else:
        row = warn_start_row + 1
        ws4.cell(row=row, column=1, value="No warnings recorded").font = Font(
            name="Calibri", size=10, italic=True, color="28A745"
        )

    # Retry metadata if available
    retry_metadata = result_data.get("retry_metadata", {})
    if retry_metadata:
        retry_row = warn_start_row + max(len(warnings), 1) + 2
        ws4.cell(row=retry_row, column=1, value="Retry Information").font = SUBHEADER_FONT
        ws4.cell(row=retry_row, column=1).fill = SUBHEADER_FILL

        ws4.cell(row=retry_row + 1, column=1, value=f"Retry Count: {retry_metadata.get('retry_count', 0)}").font = DATA_FONT
        ws4.cell(row=retry_row + 2, column=1, value=f"Retried Stages: {', '.join(retry_metadata.get('retried_stages', []))}").font = DATA_FONT
        ws4.cell(row=retry_row + 3, column=1, value=f"Last Retry: {retry_metadata.get('last_retry_at', 'N/A')}").font = DATA_FONT

    _auto_column_width(ws4, 3)

    # ── Finalize ─────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Add border styling to the BOQ data cells
    logger.info(
        "[EXPORT] Excel export complete for job %s — status=%s, boq_items=%d, warnings=%d",
        job_id, status, len(boq_items), len(warnings),
    )

    return output